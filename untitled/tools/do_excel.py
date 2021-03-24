from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path
# from tools.get_data import GetData

class DoExcel:
    def get_data(self,file_name):
        wb = load_workbook(file_name)
        # sheet = wb[sheet_name]
        test_data = []
        mode = eval(ReadConfig.get_config(project_path.case_config_path, 'MODE', 'mode'))   # 读取配置文件
        for key in mode: # 遍历这个存在配置文件里面的字典
            sheet = wb[key]
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['url'] = sheet.cell(i, 2).value
                    row_data['data'] = sheet.cell(i, 3).value
                    # if sheet.cell(i, 3).value.find('${tel_1}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel_1}', str(tel))  # 替换
                    # elif sheet.cell(i, 3).value.find('${tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel+1))
                    #  else:
                    #     row_data['data'] = sheet.cell(i, 3).value
                    row_data['title'] = sheet.cell(i, 4).value
                    row_data['http_method'] = sheet.cell(i, 5).value
                    row_data['expected'] = sheet.cell(i, 6).value
                    row_data["result_type"] = sheet.cell(i,9).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    row_data['data'] = sheet.cell(case_id+1, 3).value
                    # if sheet.cell(i, 3).value.find('${tel_1}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel_1}', str(tel))  # 替换
                    # elif sheet.cell(i, 3).value.find('${tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel+1))
                    #  else:
                    #     row_data['data'] = sheet.cell(i, 3).value
                    row_data['title'] = sheet.cell(case_id+1, 4).value
                    row_data['http_method'] = sheet.cell(case_id+1, 5).value
                    row_data['expected'] = sheet.cell(case_id+1, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
        return test_data

    def write_back(self,file_name,sheet_name,i,result,TestResult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value=result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)



if __name__ == '__main__':
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)